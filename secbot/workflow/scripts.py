"""Inline Python snippets used by the phishing-email workflow template.

Why a separate module?
* The snippets are large, multi-line strings that we don't want to inline
  in :mod:`secbot.workflow.templates` (loses syntax highlighting, makes
  the template harder to diff).
* Storing them as ``str`` constants here lets us import + unit-test the
  snippets in isolation by ``exec(...)``-ing them with a fixture stdin.

The runtime contract for each snippet:

* Reads a single JSON object from stdin (the workflow runner pipes the
  template's interpolated ``stdin`` through ``python3 -``).
* Writes ONE flat JSON object to stdout — every consumer (the runner, the
  Lua plugin) parses ``stepResults.<step>.output.stdout``.
* On any internal failure: still emit a JSON object with ``error`` set so
  the next step has something to operate on. The script must NEVER print
  Python tracebacks to stdout; tracebacks go to stderr only.
* No third-party imports. ``sqlite3`` is pulled lazily and
  failures degrade gracefully (the cache and write-back layers are best
  effort — the workflow keeps running without them).

Spec: PRD §R1, §R5 (容错策略), §Technical Notes (ScriptExecutor 60s 上限).
"""

from __future__ import annotations


# ---------------------------------------------------------------------------
# step1 — 特征提取 + SQLite 去重 + 脱敏
# ---------------------------------------------------------------------------


PHISHING_STEP1_CODE = r'''
import base64
import email
import email.policy
import hashlib
import json
import os
import re
import sqlite3
import sys
import time
import traceback
from urllib.parse import urlparse


_DB_PATH = os.environ.get(
    "PHISHING_DB_PATH",
    os.path.join(os.getcwd(), "detection_results.db"),
)

_WHITELIST_PATH = os.environ.get(
    "PHISHING_WHITELIST_PATH",
    os.path.expanduser("~/.secbot/config/phishing-domain-whitelist.json"),
)


def _load_whitelist_hint() -> str:
    """Read domain whitelist config and return a prompt hint string.

    Returns an empty string when the config is missing or unreadable so
    the LLM prompt simply omits the whitelist block.
    """
    try:
        with open(_WHITELIST_PATH, "r", encoding="utf-8") as fh:
            cfg = json.load(fh)
        domains = cfg.get("domains") or []
        if not domains:
            return ""
        domain_list = ", ".join(str(d) for d in domains)
        return (
            "【官方域名白名单规则】\n"
            f"白名单后缀：{domain_list}\n"
            "当发件人域名以以上任一后缀结尾时，该邮件来自可信官方机构，"
            "confidence 必须 ≤ 0.1，suggested_action 必须为「放行」，"
            "除非存在极强的其他可疑特征。\n"
            "不要因为发件人使用了紧迫语气或包含链接就对官方域名邮件给出高置信度。"
        )
    except Exception:
        return ""


def _lookup_sqlite_cache(chash: str) -> dict | None:
    """Return the most recent LLM result for *chash* from the local SQLite DB.

    Only genuine LLM results are used (reason not starting with
    "LLM skipped").  Returns None when not found or on any DB error.
    """
    if not os.path.isfile(_DB_PATH):
        return None
    try:
        conn = sqlite3.connect(_DB_PATH, timeout=1.0)
        conn.row_factory = sqlite3.Row
        row = conn.execute(
            """
            SELECT ai_confidence, ai_reason, action, risk_factors
            FROM   detection_results
            WHERE  content_hash = ?
              AND  ai_confidence IS NOT NULL
              AND  (ai_reason IS NULL OR ai_reason NOT LIKE 'LLM skipped%')
            ORDER  BY id DESC
            LIMIT  1
            """,
            (chash,),
        ).fetchone()
        conn.close()
        if row is None:
            return None
        import json as _json
        try:
            rf = _json.loads(row["risk_factors"]) if row["risk_factors"] else []
        except Exception:
            rf = []
        return {
            "confidence": float(row["ai_confidence"] or 0.0),
            "reason": str(row["ai_reason"] or "(cached)"),
            "suggested_action": str(row["action"] or ""),
            "risk_factors": rf,
        }
    except Exception:
        return None


def _extract_plain_text(raw: str) -> str:
    """Extract plain text from raw MIME email.

    Uses Python email module to correctly parse multipart structure
    and extract text/plain (preferred) or stripped text/html.
    Falls back to regex-based HTML stripping if MIME parsing fails.
    """
    try:
        msg = email.message_from_string(raw, policy=email.policy.default)
        # Prefer text/plain, fall back to text/html
        body_part = msg.get_body(preferencelist=("plain", "html"))
        if body_part:
            content = body_part.get_content()
            if body_part.get_content_type() == "text/html":
                content = re.sub(r"<[^>]+>", " ", content)
            return content
        # Walk parts as fallback
        for part in msg.walk():
            ct = part.get_content_type()
            if ct == "text/plain":
                return part.get_content()
            if ct == "text/html":
                return re.sub(r"<[^>]+>", " ", part.get_content())
    except Exception:
        pass
    # Last resort: strip HTML from raw input
    return re.sub(r"<[^>]+>", " ", raw)


def _content_hash(sender: str, subject: str, body: str, attachments_info: str = "") -> str:
    h = hashlib.sha256()
    h.update((sender or "").strip().lower().encode("utf-8", "replace"))
    h.update(b"|")
    h.update((subject or "").strip().encode("utf-8", "replace"))
    h.update(b"|")
    # Extract real text content from MIME body, then normalize.
    plain = _extract_plain_text(body or "")
    normalized = re.sub(r"\s+", " ", plain).strip().lower()
    h.update(normalized.encode("utf-8", "replace"))
    # Include attachment features (filename + Magika label) so that
    # emails with different attachments get different hashes.
    if attachments_info:
        h.update(b"|")
        h.update(attachments_info.encode("utf-8", "replace"))
    return h.hexdigest()


_SUSP_TLD = {"zip", "top", "xyz", "click", "link", "loan", "country"}
_LOOKALIKE_TOKENS = (
    "paypa1", "app1e", "m1crosoft", "gma1l", "amaz0n", "gith0b", "secur",
)


def _suspicious_url(url: str) -> bool:
    try:
        parsed = urlparse(url)
    except Exception:
        return False
    host = (parsed.hostname or "").lower()
    if not host:
        return True
    tld = host.rsplit(".", 1)[-1] if "." in host else ""
    if tld in _SUSP_TLD:
        return True
    if any(tok in host for tok in _LOOKALIKE_TOKENS):
        return True
    if re.fullmatch(r"\d+\.\d+\.\d+\.\d+", host):
        return True
    return False


# ---------------------------------------------------------------------------
# Magika attachment analysis
# ---------------------------------------------------------------------------

_magika_instance = None
_magika_error = None


def _init_magika():
    """Lazy-init Magika singleton. Returns (instance, error_str)."""
    global _magika_instance, _magika_error
    if _magika_instance is not None:
        return _magika_instance, None
    if _magika_error is not None:
        return None, _magika_error
    try:
        from magika import Magika, PredictionMode
        _magika_instance = Magika(
            prediction_mode=PredictionMode.HIGH_CONFIDENCE
        )
        return _magika_instance, None
    except ImportError:
        _magika_error = "not_installed"
        return None, _magika_error
    except Exception as exc:
        _magika_error = f"init_failed: {exc}"
        return None, _magika_error


# Magika label -> set of expected file extensions (lowercase, no dot)
_MAGIKA_LABEL_EXTENSIONS = {
    "pebin": {"exe", "dll"},
    "elf": {"so", ""},
    "macho": {""},
    "pdf": {"pdf"},
    "doc": {"doc"},
    "docx": {"docx", "docm"},
    "xls": {"xls"},
    "xlsx": {"xlsx", "xlsm"},
    "ppt": {"ppt"},
    "pptx": {"pptx", "pptm"},
    "javascript": {"js", "mjs"},
    "powershell": {"ps1", "psm1"},
    "vba": {"vba", "bas"},
    "batch": {"bat", "cmd"},
    "shell": {"sh", "bash"},
    "python": {"py", "pyw"},
    "zip": {"zip"},
    "rar": {"rar"},
    "sevenzip": {"7z"},
    "gzip": {"gz", "gzip"},
    "tar": {"tar"},
    "html": {"html", "htm"},
    "xml": {"xml"},
    "json": {"json"},
    "csv": {"csv"},
    "txt": {"txt", "log"},
    "markdown": {"md", "markdown"},
    "yaml": {"yaml", "yml"},
    "ini": {"ini", "cfg", "conf"},
    "java": {"java"},
    "c": {"c", "h"},
    "cpp": {"cpp", "cc", "cxx", "hpp"},
    "rust": {"rs"},
    "go": {"go"},
    "ruby": {"rb"},
    "php": {"php"},
    "sql": {"sql"},
    "onnx": {"onnx"},
    "png": {"png"},
    "jpeg": {"jpg", "jpeg"},
    "gif": {"gif"},
    "bmp": {"bmp"},
    "ico": {"ico"},
    "svg": {"svg"},
    "woff": {"woff"},
    "woff2": {"woff2"},
    "ttf": {"ttf"},
    "otf": {"otf"},
    "mp3": {"mp3"},
    "mp4": {"mp4"},
    "wav": {"wav"},
    "flac": {"flac"},
    "ogg": {"ogg"},
    "webm": {"webm"},
    "webp": {"webp"},
    "epub": {"epub"},
    "rtf": {"rtf"},
    "latex": {"tex"},
    "iso": {"iso"},
    "apk": {"apk"},
    "jar": {"jar"},
    "dex": {"dex"},
    "smali": {"smali"},
    "lnk": {"lnk"},
    "torrent": {"torrent"},
    "sqlite": {"sqlite", "db"},
    "pcap": {"pcap", "cap"},
    "postscript": {"ps", "eps"},
    "eml": {"eml"},
    "mht": {"mht"},
    "wasm": {"wasm"},
}

# Extensions that indicate macro-capable files
_MACRO_OLD_EXTENSIONS = {"doc", "xls", "ppt"}
_MACRO_EXPLICIT_EXTENSIONS = {"docm", "xlsm", "pptm", "dot", "dotm"}


def _get_extension(filename: str) -> str:
    """Extract lowercase extension without dot from filename."""
    if not filename:
        return ""
    dot_idx = filename.rfind(".")
    if dot_idx < 0 or dot_idx == len(filename) - 1:
        return ""
    return filename[dot_idx + 1:].lower()


def _check_extension_mismatch(filename: str, magika_label: str) -> bool:
    """Check if declared extension is inconsistent with Magika label."""
    declared_ext = _get_extension(filename)
    if not declared_ext or not magika_label:
        return False
    expected_exts = _MAGIKA_LABEL_EXTENSIONS.get(magika_label)
    if expected_exts is None:
        return False
    return declared_ext not in expected_exts


def _check_macro_capable(filename: str, magika_label: str) -> bool:
    """Check if file is macro-capable (old format or explicit macro file)."""
    declared_ext = _get_extension(filename)
    if declared_ext in _MACRO_OLD_EXTENSIONS:
        return True
    if declared_ext in _MACRO_EXPLICIT_EXTENSIONS:
        return True
    if magika_label in ("doc", "xls", "ppt"):
        return True
    return False


def _analyze_attachments(attachments_raw):
    """Analyze attachments with Magika. Returns (results_list, attachments_info_str)."""
    if not attachments_raw:
        return [], ""

    if isinstance(attachments_raw, str):
        try:
            attachments = json.loads(attachments_raw or "[]")
        except Exception:
            attachments = []
    elif isinstance(attachments_raw, list):
        attachments = attachments_raw
    else:
        attachments = []

    if not attachments:
        return [], ""

    magika, magika_err = _init_magika()

    results = []
    info_pairs = []

    for att in attachments:
        if not isinstance(att, dict):
            continue
        filename = str(att.get("filename") or "unknown")
        content_type = str(att.get("content_type") or "")
        content_b64 = str(att.get("content_base64") or "")
        original_size = int(att.get("original_size") or 0)

        result = {
            "filename": filename,
            "content_type": content_type,
            "original_size": original_size,
            "magika_label": "",
            "magika_score": 0.0,
            "extension_mismatch": False,
            "is_macro_capable": False,
        }

        if magika_err:
            result["magika_error"] = magika_err
        elif not content_b64:
            result["magika_error"] = "empty_content"
        else:
            try:
                raw_bytes = base64.b64decode(content_b64)
                mr = magika.identify_bytes(raw_bytes)
                label = str(mr.output.label)
                score = float(mr.score)
                result["magika_label"] = label
                result["magika_score"] = round(score, 4)
                result["extension_mismatch"] = _check_extension_mismatch(
                    filename, label
                )
                result["is_macro_capable"] = _check_macro_capable(
                    filename, label
                )
            except Exception as exc:
                result["magika_error"] = f"identify_failed: {exc}"

        results.append(result)
        info_pairs.append(f"{filename}|{result['magika_label']}")

    info_pairs.sort()
    attachments_info = ";".join(info_pairs)

    return results, attachments_info


def _emit(payload: dict) -> None:
    sys.stdout.write(json.dumps(payload, ensure_ascii=False))
    sys.stdout.write("\n")


def main() -> int:
    raw = sys.stdin.read()
    try:
        data = json.loads(raw or "{}")
    except Exception as exc:
        _emit({
            "error": f"step1.input_parse: {exc}",
            "cache_hit": False,
            "features": {},
        })
        return 0

    sender = str(data.get("sender") or "")
    subject = str(data.get("subject") or "")
    body = str(data.get("body") or "")
    rspamd_score_raw = str(data.get("rspamd_score") or "0")
    urls_in = data.get("urls") or []
    if isinstance(urls_in, str):
        try:
            urls_in = json.loads(urls_in or "[]")
        except Exception:
            urls_in = []
    if not isinstance(urls_in, list):
        urls_in = []
    urls = [str(u) for u in urls_in[:50]]

    # Parse attachments (dual-shape: JSON string or list, same as urls)
    attachments_raw = data.get("attachments") or "[]"

    # Analyze attachments with Magika
    attachment_results, attachments_info = _analyze_attachments(attachments_raw)

    sender_local, _, sender_domain = sender.partition("@")
    body_excerpt = re.sub(r"\s+", " ", body).strip()[:600]
    suspicious_domains = sorted({
        urlparse(u).hostname or ""
        for u in urls
        if _suspicious_url(u)
    } - {""})

    chash = _content_hash(sender, subject, body, attachments_info)

    cache_hit = False
    cached_result = None

    # SQLite-based cache lookup (always available, no extra deps)
    cached_result = _lookup_sqlite_cache(chash)
    if cached_result is not None:
        cache_hit = True

    try:
        rspamd_score = float(rspamd_score_raw)
    except Exception:
        rspamd_score = 0.0

    whitelist_hint = _load_whitelist_hint()

    # Build attachment summary for LLM prompt
    attachment_count = len(attachment_results)
    macro_count = sum(1 for a in attachment_results if a.get("is_macro_capable"))
    mismatch_count = sum(1 for a in attachment_results if a.get("extension_mismatch"))
    att_summary_parts = []
    for a in attachment_results:
        parts = [f"{a['filename']}({a.get('magika_label') or 'unknown'})"]
        if a.get("extension_mismatch"):
            parts.append("扩展名不匹配!")
        if a.get("is_macro_capable"):
            parts.append("宏文件")
        if a.get("magika_error"):
            parts.append(f"错误:{a['magika_error']}")
        att_summary_parts.append(" ".join(parts))
    attachment_summary = "; ".join(att_summary_parts) if att_summary_parts else "无附件"

    _emit({
        "cache_hit": cache_hit,
        "cached_result": cached_result,
        "content_hash": chash,
        "rspamd_score": rspamd_score,
        "workflow_start_ms": int(time.time() * 1000),
        "features": {
            "sender_full": sender,
            "sender_local": sender_local,
            "sender_domain": sender_domain,
            "subject": subject[:200],
            "body_excerpt": body_excerpt,
            "url_count": len(urls),
            "suspicious_domains": suspicious_domains[:10],
            "recipient": str(data.get("recipient") or ""),
            "domain_whitelist_hint": whitelist_hint,
            "attachment_count": attachment_count,
            "attachment_summary": attachment_summary,
            "attachment_details": attachment_results,
            "macro_capable_count": macro_count,
            "extension_mismatch_count": mismatch_count,
        },
    })
    return 0


try:
    sys.exit(main())
except Exception:
    sys.stderr.write(traceback.format_exc())
    # Defensive last-resort payload so step3 always has structure to read.
    sys.stdout.write(json.dumps({
        "error": "step1.unhandled",
        "cache_hit": False,
        "features": {},
    }, ensure_ascii=False) + "\n")
    sys.exit(0)
'''


# ---------------------------------------------------------------------------
# step3 — 聚合 + add_score 计算 + 回写业务 SQLite
# ---------------------------------------------------------------------------


PHISHING_STEP3_CODE = r'''
import json
import os
import sqlite3
import sys
import time
import traceback


_DB_PATH = os.environ.get(
    "PHISHING_DB_PATH",
    os.path.join(os.getcwd(), "detection_results.db"),
)


def _migrate_drop_unique(conn: sqlite3.Connection) -> None:
    """Drop UNIQUE on content_hash so repeated detections all get stored.

    SQLite has no DROP CONSTRAINT; the only way is to recreate the table.
    Safe to call repeatedly — checks the schema string first.
    """
    try:
        row = conn.execute(
            "SELECT sql FROM sqlite_master"
            " WHERE type='table' AND name='detection_results'"
        ).fetchone()
        if row is None or "content_hash TEXT UNIQUE" not in (row[0] or ""):
            return  # nothing to do
        cols = [
            r[1]
            for r in conn.execute(
                "PRAGMA table_info(detection_results)"
            ).fetchall()
        ]
        cols_csv = ", ".join(cols)
        new_sql = (
            row[0]
            .replace("detection_results", "detection_results_new", 1)
            .replace("content_hash TEXT UNIQUE", "content_hash TEXT")
        )
        conn.execute(new_sql)
        conn.execute(
            f"INSERT INTO detection_results_new ({cols_csv})"
            f" SELECT {cols_csv} FROM detection_results"
        )
        conn.execute("DROP TABLE detection_results")
        conn.execute(
            "ALTER TABLE detection_results_new RENAME TO detection_results"
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_detection_created_at"
            " ON detection_results(created_at)"
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_detection_content_hash"
            " ON detection_results(content_hash)"
        )
        conn.commit()
    except Exception as exc:
        sys.stderr.write(f"[step3] migrate_drop_unique failed: {exc}\n")


def _ensure_table(conn: sqlite3.Connection) -> None:
    # Drop UNIQUE on content_hash first so the following CREATE TABLE
    # definition (content_hash TEXT NOT NULL, no UNIQUE) is the canonical one.
    _migrate_drop_unique(conn)
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS detection_results (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            content_hash TEXT NOT NULL,
            sender TEXT,
            subject TEXT,
            ai_confidence REAL,
            ai_reason TEXT,
            action TEXT,
            created_at TEXT,
            processed_time_ms INTEGER,
            risk_factors TEXT,
            rspamd_score REAL,
            final_score REAL,
            rspamd_action TEXT
        )
        """
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_detection_created_at "
        "ON detection_results(created_at)"
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_detection_content_hash "
        "ON detection_results(content_hash)"
    )
    # --- Migrations for existing tables ---
    # Add new columns
    for col, typ in [
        ("risk_factors", "TEXT"),
        ("rspamd_score", "REAL"),
        ("final_score", "REAL"),
        ("rspamd_action", "TEXT"),
        ("attachments_json", "TEXT"),
    ]:
        try:
            conn.execute(
                f"ALTER TABLE detection_results ADD COLUMN {col} {typ}"
            )
        except Exception:
            pass
    # Rename ai_suspicion_level -> ai_confidence (legacy migration)
    try:
        conn.execute(
            "ALTER TABLE detection_results RENAME COLUMN ai_suspicion_level"
            " TO ai_confidence"
        )
    except Exception:
        pass


def _display_action(score) -> str:
    """Derive the dashboard action label from the final score.

    Keeps the displayed verdict consistent with rspamd's real behaviour
    (the score thresholds), instead of the LLM's subjective
    ``suggested_action`` which may say "标记" even at a low score.
      >= 15 → 拒绝 (reject)
      >=  6 → 标记 (add_header / mark as spam)
      <   6 → 放行 (normal delivery)
    """
    try:
        s = float(score)
    except (TypeError, ValueError):
        return "放行"
    if s >= 15:
        return "拒绝"
    if s >= 6:
        return "标记"
    return "放行"


def _persist_sqlite(row: dict) -> bool:
    db_dir = os.path.dirname(_DB_PATH)
    if db_dir:
        try:
            os.makedirs(db_dir, exist_ok=True)
        except Exception:
            pass
    try:
        conn = sqlite3.connect(_DB_PATH, timeout=2.0)
    except Exception as exc:
        sys.stderr.write(f"[step3] sqlite connect failed: {exc}\n")
        return False
    try:
        _ensure_table(conn)
        conn.execute(
            """
            INSERT INTO detection_results
                (content_hash, sender, subject, ai_confidence,
                 ai_reason, action, created_at, processed_time_ms,
                 risk_factors, rspamd_score, final_score, rspamd_action,
                 attachments_json)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                row.get("content_hash"),
                row.get("sender"),
                row.get("subject"),
                float(row.get("confidence") or 0.0),
                row.get("reason"),
                _display_action(row.get("final_score")),
                row.get("created_at"),
                int(row.get("processed_time_ms") or 0),
                json.dumps(row.get("risk_factors") or [], ensure_ascii=False),
                float(row.get("rspamd_score") or 0.0) if row.get("rspamd_score") is not None else None,
                float(row.get("final_score") or 0.0) if row.get("final_score") is not None else None,
                row.get("rspamd_action") or "",
                row.get("attachments_json") or "",
            ),
        )
        conn.commit()
        return True
    except Exception as exc:
        sys.stderr.write(f"[step3] sqlite persist failed: {exc}\n")
        return False
    finally:
        try:
            conn.close()
        except Exception:
            pass


def _add_score_for(suspicion_level: float) -> float:
    """Map LLM suspicion level to rspamd score delta.

    Formula:  add_score = suspicion_level × 8.0

    Suspicion level (0.0–1.0) represents how suspicious the email is.
    A fixed multiplier keeps the scoring simple and predictable:
      - 1.0 → 8.0 (max addition)
      - 0.6 → 4.8
      - 0.2 → 1.6
      - 0.0 → 0.0
    """
    return round(suspicion_level * 8.0, 2)


def _emit(payload: dict) -> None:
    sys.stdout.write(json.dumps(payload, ensure_ascii=False))
    sys.stdout.write("\n")


def _rspamd_action(score: float) -> str:
    """Derive the rspamd final action from total score.

    Thresholds (from rspamadm configdump):
      >= 15  → reject (拒绝投递，直接退信)
      >=  6  → add_header (投递，标记为 spam)
      >=  4  → greylist (临时拒绝，要求重试)
      <   4  → accept (正常投递)
    """
    if score >= 15:
        return "reject"
    if score >= 6:
        return "add_header"
    if score >= 4:
        return "greylist"
    return "accept"


def main() -> int:
    started = time.time()
    raw = sys.stdin.read()
    try:
        data = json.loads(raw or "{}")
    except Exception as exc:
        _emit({
            "error": f"step3.input_parse: {exc}",
            "add_score": 0.0,
            "confidence": 0.0,
            "reason": "step3 input parse failed; defaulting add_score=0",
            "suggested_action": "放行",
            "risk_factors": [],
        })
        return 0

    step1 = data.get("step1") or {}
    step2 = data.get("step2") or {}
    rspamd_score = data.get("rspamd_score")

    features = step1.get("features") or {}
    chash = step1.get("content_hash") or ""
    sender = features.get("sender_full") or ""
    subject = features.get("subject") or ""
    # Use workflow_start_ms from step1 to calculate total processing time
    workflow_start_ms = step1.get("workflow_start_ms") or 0

    cache_hit = bool(step1.get("cache_hit"))
    cached = step1.get("cached_result") or {}

    if cache_hit and cached:
        # Trust the cached judgement verbatim; recompute add_score so
        # tuning the matrix takes effect on next read without bumping TTL.
        confidence = float(cached.get("confidence") or 0.0)
        reason = str(cached.get("reason") or "(cached)")
        suggested_action = str(cached.get("suggested_action") or "")
        risk_factors = list(cached.get("risk_factors") or [])
        add_score = float(cached.get("add_score", _add_score_for(confidence)))
    else:
        # Accept two shapes for ``step2``:
        # 1. The full LlmExecutor wrapper ``{content, parsed, ...}`` --
        #    business JSON nested under ``.parsed``.
        # 2. The already-unwrapped business dict (``{confidence,
        #    reason, ...}`` at the top level), which is what the
        #    current phishing template's stdin produces by interpolating
        #    the step2 parsed result directly.
        # NOTE: do NOT write the ``$``+``{...}`` placeholder syntax here --
        # this comment lives inside ``args.code`` which the runner runs
        # through ``interpolate`` before exec, so the literal would be
        # substituted with the real (URL-containing) JSON and trip the
        # SSRF guard in ``ExecTool._guard_command``.
        # Detection is purely structural: if ``parsed`` is a dict, use
        # it; else if ``step2`` itself has the business keys, use it.
        parsed: dict | None = None
        if isinstance(step2, dict):
            inner = step2.get("parsed")
            if isinstance(inner, dict):
                parsed = inner
            elif "confidence" in step2 or "reason" in step2:
                parsed = step2
        if not isinstance(parsed, dict):
            # LLM was skipped (rspamd_score outside [4,10]) or errored.
            # Default to放行 with a clear reason.
            early_now_ms = int(time.time() * 1000)
            early_processed = (early_now_ms - workflow_start_ms) if workflow_start_ms > 0 else int((time.time() - started) * 1000)
            early_created_at = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
            early_result = {
                "add_score": 0.0,
                "confidence": 0.0,
                "reason": "LLM skipped or unavailable; default add_score=0",
                "suggested_action": "放行",
                "risk_factors": [],
                "content_hash": chash,
                "sender": sender,
                "subject": subject,
                "rspamd_score": rspamd_score,
                "final_score": float(rspamd_score or 0.0),
                "rspamd_action": _rspamd_action(float(rspamd_score or 0.0)),
                "processed_time_ms": early_processed,
                "created_at": early_created_at,
                "attachments_json": json.dumps(
                    features.get("attachment_details") or [],
                    ensure_ascii=False
                ),
            }
            # Persist to SQLite even for LLM-skipped path
            if chash:
                _persist_sqlite(early_result)
            _emit(early_result)
            return 0

        confidence = float(parsed.get("confidence") or 0.0)
        reason = str(parsed.get("reason") or "")
        suggested_action = str(parsed.get("suggested_action") or "")
        risk_factors = list(parsed.get("risk_factors") or [])
        add_score = _add_score_for(confidence)

    # --- Attachment scoring (macro-capable bonus) ---
    # PRD 06-01-magika-attachment-detect §R9: each macro-capable file
    # adds +2.0 to add_score, regardless of LLM judgement.
    attachment_details = features.get("attachment_details") or []
    macro_capable_count = int(features.get("macro_capable_count") or 0)
    mismatch_count = int(features.get("extension_mismatch_count") or 0)
    macro_bonus = macro_capable_count * 2.0
    if macro_bonus > 0:
        add_score = round(float(add_score) + macro_bonus, 2)
        risk_factors.append(
            f"含 {macro_capable_count} 个宏能力附件 (+{macro_bonus:.1f})"
        )
    if mismatch_count > 0:
        mismatch_names = [
            a.get("filename", "?")
            for a in attachment_details
            if a.get("extension_mismatch")
        ]
        risk_factors.append(
            f"附件扩展名不匹配: {', '.join(mismatch_names[:3])}"
        )
    attachments_json = json.dumps(attachment_details, ensure_ascii=False)

    # Calculate total workflow processing time (from step1 start to now)
    now_ms = int(time.time() * 1000)
    if workflow_start_ms > 0:
        processed_ms = now_ms - workflow_start_ms
    else:
        processed_ms = int((time.time() - started) * 1000)
    created_at = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())

    rspamd_score_f = float(rspamd_score or 0.0)
    final_score = round(rspamd_score_f + float(add_score), 2)
    rspamd_action = _rspamd_action(final_score)

    result = {
        "add_score": float(add_score),
        "confidence": float(confidence),
        "reason": reason,
        "suggested_action": suggested_action,
        "risk_factors": risk_factors,
        "content_hash": chash,
        "sender": sender,
        "subject": subject,
        "rspamd_score": rspamd_score_f,
        "final_score": final_score,
        "rspamd_action": rspamd_action,
        "processed_time_ms": processed_ms,
        "created_at": created_at,
        "attachments_json": attachments_json,
    }

    # Always persist to SQLite for dashboard visibility.
    if chash:
        _persist_sqlite(result)

    _emit(result)
    return 0


try:
    sys.exit(main())
except Exception:
    sys.stderr.write(traceback.format_exc())
    sys.stdout.write(json.dumps({
        "error": "step3.unhandled",
        "add_score": 0.0,
        "confidence": 0.0,
        "reason": "step3 unhandled exception; defaulting add_score=0",
        "suggested_action": "放行",
        "risk_factors": [],
    }, ensure_ascii=False) + "\n")
    sys.exit(0)
'''


# ---------------------------------------------------------------------------
# Log analysis — step1: 纯读取文件内容，不做解析，交给 LLM
# ---------------------------------------------------------------------------


LOG_ANALYSIS_STEP1_CODE = r'''
import json
import os
import sys
import traceback


def _read_xlsx_as_text(file_path: str) -> str:
    """Read xlsx file from disk, convert all cells to tab-separated text."""
    try:
        import openpyxl
    except ImportError:
        return "ERROR: openpyxl not available, cannot read .xlsx files"

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    except Exception as exc:
        return f"ERROR: Failed to open xlsx: {exc}"

    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        parts.append(f"[Sheet: {sheet_name}]")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            parts.append("\t".join(cells))
    wb.close()
    return "\n".join(parts)


def main() -> int:
    raw = sys.stdin.read()
    try:
        data = json.loads(raw or "{}")
    except Exception:
        # Malformed input → emit safe empty payload so downstream
        # (LLM) has something to work with.
        sys.stdout.write(json.dumps({
            "log_content": "",
            "file_name": "unknown",
            "log_format": "unknown",
            "char_count": 0,
        }))
        return 0

    log_path = str(data.get("log_path") or "").strip()

    if not log_path:
        # Upload mode — content flows directly from inputs.log_content
        # into the LLM prompt.  This step just passes through empty
        # content so the template interpolation in step2 (LLM) always
        # sees a valid string.
        sys.stdout.write(json.dumps({
            "log_content": "",
            "file_name": "uploaded",
            "log_format": "uploaded_text",
            "char_count": 0,
        }))
        return 0

    # -------------------------------------------------------------------
    # Path mode — read the file from the server filesystem
    # -------------------------------------------------------------------
    file_name = os.path.basename(log_path)
    lower = log_path.lower()

    # Format detection
    if lower.endswith((".xlsx", ".xlsm", ".xls")):
        fmt = "xlsx"
        try:
            content = _read_xlsx_as_text(log_path)
        except Exception as exc:
            sys.stdout.write(json.dumps({
                "log_content": "",
                "file_name": file_name,
                "log_format": fmt,
                "char_count": 0,
                "error": f"xlsx_read: {exc}",
            }))
            return 0
    else:
        if lower.endswith(".csv"):
            fmt = "csv"
        elif lower.endswith((".tsv", ".tab")):
            fmt = "tsv"
        else:
            fmt = "txt"
        try:
            with open(log_path, "r", encoding="utf-8", errors="replace") as f:
                content = f.read()
        except Exception as exc:
            sys.stdout.write(json.dumps({
                "log_content": "",
                "file_name": file_name,
                "log_format": fmt,
                "char_count": 0,
                "error": f"file_read: {exc}",
            }))
            return 0

    # Truncate to stay within the exec tool _MAX_OUTPUT (100 KB).  The
    # JSON wrapper adds ~15 % overhead; 50 000 chars is a safe ceiling.
    max_chars = 50000
    char_count = len(content)
    total_entries = sum(1 for line in content.splitlines() if line.strip())
    if char_count > max_chars:
        content = content[:max_chars]

    sys.stdout.write(json.dumps({
        "log_content": content,
        "file_name": file_name,
        "log_format": fmt,
        "char_count": char_count,
        "total_entries": total_entries,
    }, ensure_ascii=False))
    return 0


try:
    sys.exit(main())
except Exception:
    sys.stderr.write(traceback.format_exc())
    sys.exit(0)
'''


# ---------------------------------------------------------------------------
# Log analysis — step3: LLM 结果入库 + 报告生成
# ---------------------------------------------------------------------------


LOG_ANALYSIS_STEP3_CODE = r'''
import json
import os
import sqlite3
import sys
import time
import traceback


_DB_PATH = os.environ.get(
    "LOG_ANALYSIS_DB_PATH",
    os.path.join(os.getcwd(), "detection_results.db"),
)


def _ensure_table(conn: sqlite3.Connection) -> None:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS log_analysis (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            workflow_run_id TEXT,
            file_name TEXT,
            log_format TEXT,
            char_count INTEGER,
            analysis_timestamp TEXT,
            anomaly_count INTEGER,
            critical_count INTEGER,
            high_count INTEGER,
            medium_count INTEGER,
            low_count INTEGER,
            summary TEXT,
            analysis_json TEXT,
            created_at TEXT
        )
        """
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_log_analysis_created_at "
        "ON log_analysis(created_at)"
    )
    for col, typ in [
        ("critical_count", "INTEGER"),
        ("high_count", "INTEGER"),
        ("medium_count", "INTEGER"),
        ("low_count", "INTEGER"),
        ("char_count", "INTEGER"),
        ("total_entries", "INTEGER"),
    ]:
        try:
            conn.execute(f"ALTER TABLE log_analysis ADD COLUMN {col} {typ}")
        except Exception:
            pass


def _persist(conn: sqlite3.Connection, row: dict) -> int:
    """Persist a log-analysis row and return the new rowid (0 on failure)."""
    try:
        _ensure_table(conn)
        cursor = conn.execute(
            """
            INSERT INTO log_analysis
                (workflow_run_id, file_name, log_format, char_count,
                 analysis_timestamp, anomaly_count,
                 critical_count, high_count, medium_count, low_count,
                 total_entries,
                 summary, analysis_json, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                row.get("workflow_run_id"),
                row.get("file_name"),
                row.get("log_format"),
                int(row.get("char_count") or 0),
                row.get("analysis_timestamp"),
                int(row.get("anomaly_count") or 0),
                int(row.get("critical_count") or 0),
                int(row.get("high_count") or 0),
                int(row.get("medium_count") or 0),
                int(row.get("low_count") or 0),
                int(row.get("total_entries") or 0),
                row.get("summary"),
                row.get("analysis_json"),
                row.get("created_at"),
            ),
        )
        conn.commit()
        return cursor.lastrowid or 0
    except Exception as exc:
        sys.stderr.write(f"[log-analysis-step3] persist failed: {exc}\n")
        return 0


def _emit(payload: dict) -> None:
    sys.stdout.write(json.dumps(payload, ensure_ascii=False))
    sys.stdout.write("\n")


def main() -> int:
    raw = sys.stdin.read()
    try:
        data = json.loads(raw or "{}")
    except Exception as exc:
        _emit({"error": f"step3.input_parse: {exc}", "report": "", "summary": ""})
        return 0

    # ── file_name ──
    # Path mode: step1 extracts the basename from the path (e.g.
    # "/var/log/auth.log" → "auth.log"); prefer it over the input
    # default "unknown.log".
    step1 = data.get("step1") or {}
    if isinstance(step1, dict):
        step1_name = str(step1.get("file_name") or "")
        if step1_name:
            file_name = step1_name
        else:
            file_name = str(data.get("file_name") or "")
    else:
        file_name = str(data.get("file_name") or "")
    if not file_name:
        file_name = "unknown.log"

    # ── step1 metrics (available in path mode) ──
    step1_valid = isinstance(step1, dict) and bool(step1)
    log_format = str(step1.get("log_format") or "unknown") if step1_valid else "unknown"
    char_count = int(step1.get("char_count") or 0) if step1_valid else 0
    total_entries = int(step1.get("total_entries") or 0) if step1_valid else 0

    # ── Upload-mode fallback ──
    # When step1 is skipped (upload mode), derive format from file_name
    # extension and compute char_count from the actual log_content length
    # passed via the stdin template.
    upload_content = str(data.get("log_content") or "")
    if not step1_valid or log_format == "unknown" or char_count == 0:
        if upload_content and char_count == 0:
            char_count = len(upload_content)
        lower_name = file_name.lower()
        if log_format in ("unknown", "uploaded_text"):
            if lower_name.endswith((".xlsx", ".xlsm", ".xls")):
                log_format = "xlsx"
            elif lower_name.endswith(".csv"):
                log_format = "csv"
            elif lower_name.endswith((".tsv", ".tab")):
                log_format = "tsv"
            elif lower_name.endswith(".json"):
                log_format = "json"
            elif lower_name.endswith(".xml"):
                log_format = "xml"
            elif lower_name.endswith((".log", ".txt")):
                log_format = "txt"
            elif upload_content:
                log_format = "txt"

    # Compute total_entries if not yet available (upload mode)
    if total_entries == 0:
        content_for_count = upload_content or (step1.get("log_content", "") if step1_valid else "")
        if content_for_count:
            total_entries = sum(1 for line in content_for_count.splitlines() if line.strip())

    # ── step2 = LLM judgement ──
    step2 = data.get("step2") or {}
    step2_parsed = step2 if isinstance(step2, dict) else {}
    # Handle {parsed: {...}} wrapper or raw dict
    if isinstance(step2_parsed, dict) and "parsed" in step2_parsed:
        inner = step2_parsed.get("parsed")
        if isinstance(inner, dict):
            step2_parsed = inner

    created_at = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())

    # Extract LLM result — all fields come from LLM now
    llm_confidence = 0.0
    llm_reason = ""
    llm_risk_factors = []
    llm_suggested_action = ""
    anomaly_count = 0
    anomaly_entries = []
    sev_counts = {"critical": 0, "high": 0, "medium": 0, "low": 0}

    if isinstance(step2_parsed, dict):
        llm_confidence = float(step2_parsed.get("confidence") or 0.0)
        llm_reason = str(step2_parsed.get("reason") or "")
        llm_risk_factors = list(step2_parsed.get("risk_factors") or [])
        llm_suggested_action = str(step2_parsed.get("suggested_action") or "")
        anomaly_entries = list(step2_parsed.get("anomaly_entries") or [])
        anomaly_count = int(step2_parsed.get("anomaly_count") or len(anomaly_entries))
        sev_counts = {
            "critical": int((step2_parsed.get("severity_distribution") or {}).get("critical", 0)),
            "high": int((step2_parsed.get("severity_distribution") or {}).get("high", 0)),
            "medium": int((step2_parsed.get("severity_distribution") or {}).get("medium", 0)),
            "low": int((step2_parsed.get("severity_distribution") or {}).get("low", 0)),
        }
    elif isinstance(step2_parsed, str):
        llm_reason = step2_parsed[:2000]

    # Build text report
    report_lines = [
        "=" * 60,
        "  日志分析报告",
        "=" * 60,
        f"文件名：{file_name}",
        f"格式：{log_format}",
        f"内容大小：{char_count} 字符",
        f"已检测总条目：{total_entries}",
        f"分析时间：{created_at}",
        "",
        f"【LLM 分析结论】",
        f"整体威胁置信度：{llm_confidence:.2f}",
        f"分析依据：{llm_reason}",
        f"建议处理：{llm_suggested_action}",
        f"异常条目数：{anomaly_count}",
        "",
        f"【严重级别分布】",
        f"  Critical: {sev_counts['critical']}",
        f"  High: {sev_counts['high']}",
        f"  Medium: {sev_counts['medium']}",
        f"  Low: {sev_counts['low']}",
    ]
    if llm_risk_factors:
        report_lines.append("")
        report_lines.append("风险因素：")
        for i, rf in enumerate(llm_risk_factors, 1):
            report_lines.append(f"  {i}. {rf}")
    if anomaly_entries:
        report_lines.append("")
        report_lines.append("异常详情（前20条）：")
        for i, ae in enumerate(anomaly_entries[:20], 1):
            report_lines.append(f"  {i}. {json.dumps(ae, ensure_ascii=False)[:300]}")
    report_lines.append("=" * 60)
    report = "\n".join(report_lines)

    analysis_payload = {
        "confidence": llm_confidence,
        "reason": llm_reason,
        "suggested_action": llm_suggested_action,
        "risk_factors": llm_risk_factors,
        "anomaly_entries": anomaly_entries[:50],
        "anomaly_count": anomaly_count,
        "total_entries": total_entries,
        "severity_distribution": sev_counts,
    }

    # Persist
    db_dir = os.path.dirname(_DB_PATH)
    if db_dir:
        try:
            os.makedirs(db_dir, exist_ok=True)
        except Exception:
            pass

    persisted = False
    last_id = 0
    try:
        conn = sqlite3.connect(_DB_PATH, timeout=2.0)
        last_id = _persist(conn, {
            "workflow_run_id": "",
            "file_name": file_name,
            "log_format": log_format,
            "char_count": char_count,
            "analysis_timestamp": created_at,
            "anomaly_count": anomaly_count,
            "critical_count": sev_counts["critical"],
            "high_count": sev_counts["high"],
            "medium_count": sev_counts["medium"],
            "low_count": sev_counts["low"],
            "total_entries": total_entries,
            "summary": report[:500],
            "analysis_json": json.dumps(analysis_payload, ensure_ascii=False),
            "created_at": created_at,
        })
        persisted = last_id > 0
        conn.close()
    except Exception as exc:
        sys.stderr.write(f"[log-analysis-step3] db error: {exc}\n")

    _emit({
        "report": report,
        "summary": report[:200],
        "anomaly_count": anomaly_count,
        "confidence": llm_confidence,
        "reason": llm_reason,
        "suggested_action": llm_suggested_action,
        "risk_factors": llm_risk_factors,
        "severity_distribution": sev_counts,
        "file_name": file_name,
        "char_count": char_count,
        "persisted": persisted,
        "last_id": last_id,
    })
    return 0


try:
    sys.exit(main())
except Exception:
    sys.stderr.write(traceback.format_exc())
    sys.stdout.write(json.dumps({
        "error": "step3.unhandled_exception",
        "report": "",
        "summary": "",
        "anomaly_count": 0,
    }, ensure_ascii=False) + "\n")
    sys.exit(0)
'''


__all__ = [
    "PHISHING_STEP1_CODE",
    "PHISHING_STEP3_CODE",
    "LOG_ANALYSIS_STEP1_CODE",
    "LOG_ANALYSIS_STEP3_CODE",
]
